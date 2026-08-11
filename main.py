"""PDFMasry API — secure edition (patched 2026-08-03).

Adds three Day-1-audit fixes on top of the 2026-07-30 UUID+HMAC baseline:

  1. MIME magic-byte validation before running any subprocess.
     Wrong file type now returns 415 (or 400 for empty/corrupt) instead
     of leaking 500 from the underlying tool.

  2. Per-IP rate limiting via slowapi.
     Convert endpoints: 20/minute. Downloads: 60/minute.
     Global default: 120/hour.

  3. HEAD method on /api/download/{job_id}/{token}.
     Same headers as GET, no body. Fixes crawler / link-preview probes
     that previously got 405.

Every convert endpoint still:
  * stores files in <uploads>/<job_uuid>/  (per-job isolation)
  * returns a signed one-time download URL: /api/download/<job_id>/<token>
  * token = HMAC(secret, "<job_id>.<expiry>") with 1h TTL
  * uses the ORIGINAL filename only in Content-Disposition (RFC 5987)
  * emits Cache-Control: private, no-store on downloads
  * deletes the whole job dir after 1 hour via background task

Endpoints unchanged from the frontend's perspective:
  POST /api/{tool}     → 200 {"status":"success","download_url":"/api/download/..."}
  GET  /api/download/{job_id}/{token}   → file with proper headers
  HEAD /api/download/{job_id}/{token}   → same headers, no body
"""
from __future__ import annotations

import asyncio
import hashlib
import hmac
import os
import re
import secrets
import shutil
import subprocess
import time
import unicodedata
import uuid
from urllib.parse import quote

from fastapi import (
    BackgroundTasks,
    FastAPI,
    File,
    Form,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address

# --- استدعاء مكتبات أدوبي الحديثة (V4) ---
from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
from adobe.pdfservices.operation.pdf_services import PDFServices
from adobe.pdfservices.operation.pdf_services_media_type import PDFServicesMediaType
from adobe.pdfservices.operation.pdfjobs.jobs.export_pdf_job import ExportPDFJob
from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_params import ExportPDFParams
from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_target_format import ExportPDFTargetFormat
from adobe.pdfservices.operation.pdfjobs.result.export_pdf_result import ExportPDFResult

# ───────────────────────────────────────────────────────────────────────
# Adobe usage tracker (privacy fix 2026-08-11)
# ───────────────────────────────────────────────────────────────────────
# The file cache was removed to match the site's published privacy
# promise: user files (input and output alike) are deleted within one
# hour and never retained. The previous SHA-256 cache held DOCX/XLSX
# results for 24h — a direct contradiction. See CACHE_ENABLED elsewhere
# for the flag that used to gate this; it now hard-defaults to False and
# no code path caches user file bytes.
#
# cache_stats is still imported so the admin dashboard can surface
# Adobe transaction counters. It never stores file content.
import cache_stats

# Hard-disabled. Kept as a constant only so callers can reference it
# without another rename churn. Never set to True.
CACHE_ENABLED = False

# Bump this on Adobe SDK upgrades or format-changing config changes so
# stored cache entries invalidate automatically. Format: freeform tag.
ADOBE_PROVIDER_VERSION = "adobe-sdk-v4-2026-08"

# Monthly Adobe free tier — surfaced by /api/admin/cache-stats so the
# dashboard can compute remaining quota. Bump if the paid tier is purchased.
ADOBE_MONTHLY_QUOTA = int(os.environ.get("ADOBE_MONTHLY_QUOTA", "500"))

# Admin token gates /api/admin/cache-stats. Missing token → endpoint 401s
# for everyone (fail closed, don't accidentally expose usage counters).
ADMIN_STATS_TOKEN = os.environ.get("ADMIN_STATS_TOKEN")

# ───────────────────────────────────────────────────────────────────────
# Rate limiter — per-IP quotas via X-Forwarded-For (Railway sets this)
# ───────────────────────────────────────────────────────────────────────
limiter = Limiter(key_func=get_remote_address, default_limits=["120/hour", "60/minute"])

app = FastAPI(title="PDFMasry API Complete", version="5.2-hardened")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


# ─── Structured logging + request_id middleware ─────────────────────
# Every request gets a UUID that flows into both the JSON error body
# returned to the client and the server-side log line. When a user
# reports "conversion failed on X", we can grep the log for their
# request_id and find the exact stack, without exposing internals in
# the response.
import logging
import uuid as _uuid
from starlette.middleware.base import BaseHTTPMiddleware

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] rid=%(request_id)s %(message)s",
)
_logger = logging.getLogger("pdfmasry")


class _RequestIdFilter(logging.Filter):
    """Injects request_id from the current request context if available."""
    def filter(self, record):
        rid = getattr(record, "request_id", None) or "-"
        record.request_id = rid
        return True


for h in logging.getLogger().handlers:
    h.addFilter(_RequestIdFilter())


class RequestIdMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        rid = request.headers.get("X-Request-ID") or _uuid.uuid4().hex
        request.state.request_id = rid
        try:
            response = await call_next(request)
        except HTTPException:
            raise
        except Exception as e:
            # Catch-all so uncaught exceptions never leak stack traces.
            # Detail is logged internally; client gets a stable generic message.
            _logger.exception("unhandled: %s", type(e).__name__, extra={"request_id": rid})
            return JSONResponse(
                status_code=500,
                content={
                    "detail": "حدث خطأ غير متوقع. يرجى إعادة المحاولة.",
                    "request_id": rid,
                },
                headers={"X-Request-ID": rid},
            )
        response.headers["X-Request-ID"] = rid
        return response


app.add_middleware(RequestIdMiddleware)


def _rid(request: Request) -> str:
    """Extract request_id previously injected by the middleware."""
    return getattr(request.state, "request_id", "-")


async def _run_subprocess(cmd: list[str], timeout: int = 60) -> None:
    """Run a subprocess without blocking the event loop.

    Uses asyncio.to_thread so the FastAPI worker keeps serving other
    requests while Ghostscript / qpdf / LibreOffice / pdftoppm churn.
    Enforces a per-call timeout, kills the process tree on expiry, and
    re-raises subprocess.CalledProcessError / TimeoutExpired unchanged
    so callers can pattern-match.
    """
    def _sync():
        return subprocess.run(cmd, check=True, capture_output=True, timeout=timeout)
    await asyncio.to_thread(_sync)


def _safe_http_exception(request: Request, status: int, user_detail: str, internal_reason: str = "") -> HTTPException:
    """Build an HTTPException that carries a user-safe message + request_id
    while logging the internal reason server-side only. Prefer this over
    raising raw HTTPException(detail=str(e)) — that leaks stack info.
    """
    rid = _rid(request)
    if internal_reason:
        _logger.error("%s: %s", user_detail, internal_reason, extra={"request_id": rid})
    return HTTPException(
        status_code=status,
        detail=user_detail,
        headers={"X-Request-ID": rid},
    )

# إعدادات الأمان (CORS) لحماية الباندويث الخاص بك
# CORS allow-list: env-driven so a preview / staging service can whitelist
# its Netlify Deploy Preview URL without touching production. Comma-separated
# via CORS_EXTRA_ORIGINS. The static list keeps production origins that
# don't change deploy-to-deploy.
_static_origins = [
    "https://pdfmasry.com",
    "https://www.pdfmasry.com",
    "https://taupe-rugelach-921837.netlify.app",  # historical Netlify site
    "https://pdfmasry-staging.netlify.app",       # main Netlify project
    "http://localhost:4321",
    "http://localhost:4322",
]
_extra_origins_env = os.environ.get("CORS_EXTRA_ORIGINS", "").strip()
_extra_origins = [
    o.strip() for o in _extra_origins_env.split(",") if o.strip()
] if _extra_origins_env else []
_allowed_origins = _static_origins + _extra_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

UPLOAD_DIR = "./temp_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# NOTE: A previous revision of this file called shutil.rmtree() on the
# legacy ./cache directory at boot time. Review round 2 rejected that
# pattern — auto-delete on startup, with ignore_errors=True and a path
# partially derived from PDF_CACHE_DIR, is unsafe and could wipe the
# wrong target if the env var was misconfigured. Removed. Any actual
# purge on production goes through scripts/purge_legacy_cache.py which
# is dry-run by default and requires --execute plus a hardcoded
# allowlist of paths.

# Secret used to sign download tokens. MUST be provided via env in
# production/staging — otherwise multi-replica deploys mint tokens with
# different per-boot secrets and users' download links break at random.
# The old code silently generated a random per-boot fallback; reviewer
# round 2 flagged that. We now fail loud unless APP_ENV=development.
#
# Set APP_ENV=production (or =staging) + DOWNLOAD_SECRET on Railway.
# For local dev leave APP_ENV=development (default) and the random
# per-boot value is fine because you're the only replica.
APP_ENV = os.environ.get("APP_ENV", "development").lower()
_DOWNLOAD_SECRET_ENV = os.environ.get("DOWNLOAD_SECRET")
if not _DOWNLOAD_SECRET_ENV:
    if APP_ENV in {"production", "staging"}:
        raise RuntimeError(
            "DOWNLOAD_SECRET is not set. Refusing to boot in "
            f"APP_ENV={APP_ENV!r} because a per-boot random secret "
            "would break download links across replicas or restarts."
        )
    # dev only — random per-boot is acceptable for a single local worker
    DOWNLOAD_SECRET = secrets.token_urlsafe(32)
else:
    DOWNLOAD_SECRET = _DOWNLOAD_SECRET_ENV
DOWNLOAD_TOKEN_TTL_SECONDS = 60 * 60  # 1 hour
JOB_TTL_SECONDS = 60 * 60             # cleanup after 1 hour
MAX_UPLOAD_BYTES = 50 * 1024 * 1024   # 50 MB — matches frontend claim

# ───────────────────────────────────────────────────────────────────────
# MIME validation — magic bytes only, no reliance on filename extension
# ───────────────────────────────────────────────────────────────────────
# Accepted magic byte signatures per tool family.
MAGIC_PDF = b"%PDF-"
MAGIC_ZIP = b"PK\x03\x04"          # DOCX, XLSX, PPTX (Office 2007+)
MAGIC_OLE = b"\xd0\xcf\x11\xe0"    # DOC, XLS, PPT (Office 97-2003)


def _detect_magic(path: str) -> str:
    """Return a canonical family name based on the first few bytes,
    or empty string if unrecognised. We only need to distinguish the
    types this API accepts.
    """
    try:
        with open(path, "rb") as f:
            head = f.read(8)
    except (OSError, IOError):
        return ""
    if head.startswith(MAGIC_PDF):
        return "pdf"
    if head.startswith(MAGIC_ZIP):
        return "office_zip"     # docx/xlsx/pptx — distinguish later per endpoint
    if head.startswith(MAGIC_OLE):
        return "office_ole"     # doc/xls/ppt legacy
    return ""


def _validate_input(path: str, allowed_families: set[str]) -> None:
    """Raise HTTPException if the file at `path` isn't in `allowed_families`.
    Files that don't match any known magic → 415.
    Empty files → 400.

    For Office ZIP-based files (docx/xlsx) the family match alone isn't
    enough — any ZIP starts with 'PK\x03\x04'. We call _validate_office_zip
    to inspect the archive's own membership list before accepting it,
    which blocks disguised ZIPs and ZIP bombs.
    """
    try:
        size = os.path.getsize(path)
    except OSError:
        size = 0
    if size == 0:
        raise HTTPException(status_code=400, detail="الملف فارغ")
    family = _detect_magic(path)
    if not family:
        raise HTTPException(
            status_code=415,
            detail="نوع الملف غير مدعوم — تأكد إن الملف حقيقي وسليم",
        )
    if family not in allowed_families:
        raise HTTPException(
            status_code=415,
            detail="نوع الملف لا يتوافق مع هذه الأداة",
        )
    # Deeper structural check for the ZIP-based Office formats we accept
    if family == "office_zip":
        want = None
        if "office_zip_docx" in allowed_families:
            want = "docx"
        elif "office_zip_xlsx" in allowed_families:
            want = "xlsx"
        # If the caller didn't pin a specific Office subtype we still want
        # both defenses (Zip Slip + bomb caps) applied to whichever it is.
        _validate_office_zip(path, want)


# ─── Office ZIP hardening (Phase 1 sec 2026-08-11) ─────────────────
# Limits chosen so a well-formed 50 MB DOCX/XLSX passes but common
# ZIP-bomb patterns (nested archives, extreme compression ratios) trip
# the guards long before Python's zipfile allocates gigabytes.

_MAX_ZIP_ENTRIES = 5000              # a real Office doc rarely exceeds 500
_MAX_ZIP_UNCOMPRESSED_BYTES = 500 * 1024 * 1024   # 500 MB total
_MAX_ZIP_COMPRESSION_RATIO = 200     # anything higher is almost certainly a bomb


def _validate_office_zip(path: str, want: str | None) -> None:
    """Inspect a ZIP archive to confirm it's a real DOCX/XLSX (not a
    disguised ZIP or a decompression bomb) BEFORE any downstream tool
    opens it.

    Checks:
      - archive opens as a ZIP (rules out truncated / non-ZIP disguised)
      - entry count within limit
      - no Zip Slip: no absolute paths, no '..' segments, no drive letters
      - total uncompressed size within limit
      - per-entry compression ratio within limit
      - required Office members exist for the declared type
    """
    import zipfile

    try:
        zf = zipfile.ZipFile(path, "r")
    except zipfile.BadZipFile:
        raise HTTPException(status_code=415, detail="الملف ليس Office صالحاً")

    try:
        entries = zf.infolist()
        if len(entries) > _MAX_ZIP_ENTRIES:
            raise HTTPException(status_code=413, detail="الملف يحتوي على عدد كبير جداً من العناصر")

        total_uncompressed = 0
        names = set()
        for info in entries:
            # Zip Slip defense — reject any suspicious path shape
            name = info.filename
            if not name or name.startswith("/") or name.startswith("\\"):
                raise HTTPException(status_code=415, detail="مسار مشبوه داخل الملف")
            if ".." in name.split("/") or ".." in name.split("\\"):
                raise HTTPException(status_code=415, detail="مسار مشبوه داخل الملف")
            if len(name) > 2 and name[1] == ":":  # e.g. C:\...
                raise HTTPException(status_code=415, detail="مسار مشبوه داخل الملف")

            total_uncompressed += info.file_size
            if total_uncompressed > _MAX_ZIP_UNCOMPRESSED_BYTES:
                raise HTTPException(status_code=413, detail="الحجم بعد فك الضغط يتجاوز الحد المسموح")

            if info.compress_size > 0:
                ratio = info.file_size / info.compress_size
                if ratio > _MAX_ZIP_COMPRESSION_RATIO:
                    raise HTTPException(status_code=415, detail="نسبة الضغط مشبوهة (احتمال zip bomb)")

            names.add(name)

        # Required Office members — the presence of these tells us the
        # ZIP is actually a Word/Excel document, not just any archive.
        if "[Content_Types].xml" not in names:
            raise HTTPException(status_code=415, detail="الملف ليس Office صالحاً")

        if want == "docx":
            need = {"word/document.xml"}
            if not need.issubset(names):
                raise HTTPException(status_code=415, detail="الملف ليس مستند Word صالحاً")
        elif want == "xlsx":
            need = {"xl/workbook.xml"}
            if not need.issubset(names):
                raise HTTPException(status_code=415, detail="الملف ليس مستند Excel صالحاً")
    finally:
        zf.close()


# ───────────────────────────────────────────────────────────────────────
# Security helpers — UUID job dirs, signed tokens, sanitized names
# ───────────────────────────────────────────────────────────────────────

def _new_job_id() -> str:
    """32 hex chars, 128 bits of entropy — unguessable."""
    return uuid.uuid4().hex


def _job_dir(job_id: str) -> str:
    return os.path.join(UPLOAD_DIR, job_id)


def _sign(payload: str) -> str:
    return hmac.new(
        DOWNLOAD_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def _make_token(job_id: str) -> str:
    expiry = int(time.time()) + DOWNLOAD_TOKEN_TTL_SECONDS
    sig = _sign(f"{job_id}.{expiry}")
    return f"{expiry}.{sig}"


def _verify_token(job_id: str, token: str) -> bool:
    try:
        expiry_str, sig = token.split(".", 1)
        expiry = int(expiry_str)
    except (ValueError, AttributeError):
        return False
    if time.time() > expiry:
        return False
    expected = _sign(f"{job_id}.{expiry}")
    return hmac.compare_digest(sig, expected)


def _sanitize_filename(name: str) -> str:
    """Safe display filename for Content-Disposition. Storage uses UUID."""
    name = os.path.basename(name or "file")
    name = re.sub(r"[\x00-\x1f\x7f/\\:*?\"<>|]+", "_", name)
    name = unicodedata.normalize("NFC", name).strip(" .")
    if not name:
        name = "file"
    return name[:120]


def _content_disposition(download_name: str) -> str:
    ascii_fallback = re.sub(r"[^A-Za-z0-9._-]+", "_", download_name) or "file"
    quoted = quote(download_name, safe="")
    return f'attachment; filename="{ascii_fallback}"; filename*=UTF-8\'\'{quoted}'


async def _save_upload_to_job(file: UploadFile, allowed_families: set[str]) -> tuple[str, str, str]:
    """Create a fresh job dir, save upload as input.bin, validate its magic
    bytes against `allowed_families`. On rejection the job dir is cleaned
    up and an HTTPException is raised.

    Returns (job_id, job_dir, sanitized_original_name) on success.
    """
    job_id = _new_job_id()
    jdir = _job_dir(job_id)
    os.makedirs(jdir, exist_ok=True)
    input_path = os.path.join(jdir, "input.bin")
    total = 0
    try:
        with open(input_path, "wb") as buffer:
            # Stream in chunks so we can enforce size cap without loading
            # the whole file into memory.
            while True:
                chunk = await file.read(1024 * 1024)  # 1 MB
                if not chunk:
                    break
                total += len(chunk)
                if total > MAX_UPLOAD_BYTES:
                    buffer.close()
                    shutil.rmtree(jdir, ignore_errors=True)
                    raise HTTPException(
                        status_code=413,
                        detail="حجم الملف يتجاوز الحد المسموح (50 ميجابايت)",
                    )
                buffer.write(chunk)
        _validate_input(input_path, allowed_families)
    except HTTPException:
        shutil.rmtree(jdir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(jdir, ignore_errors=True)
        _logger.exception("upload save failed: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="فشل استقبال الملف. يرجى المحاولة مرة أخرى.")

    return job_id, jdir, _sanitize_filename(file.filename or "file")


def _make_response(job_id: str, output_path: str, download_name: str, media_type: str) -> dict:
    """Move the produced file to <job_dir>/output.bin (canonical name),
    write meta so download endpoint knows the display name + type.
    """
    jdir = _job_dir(job_id)
    canonical = os.path.join(jdir, "output.bin")
    if os.path.abspath(output_path) != os.path.abspath(canonical):
        shutil.move(output_path, canonical)
    with open(os.path.join(jdir, "meta.txt"), "w", encoding="utf-8") as f:
        f.write(f"{media_type}\n{download_name}\n")
    return {
        "status": "success",
        "download_url": f"/api/download/{job_id}/{_make_token(job_id)}",
        "expires_in": DOWNLOAD_TOKEN_TTL_SECONDS,
    }


async def _delete_job_after_delay(job_id: str, delay_seconds: int = JOB_TTL_SECONDS):
    await asyncio.sleep(delay_seconds)
    jdir = _job_dir(job_id)
    if os.path.exists(jdir):
        shutil.rmtree(jdir, ignore_errors=True)

# ───────────────────────────────────────────────────────────────────────
# 1. المحرك الأساسي لأدوبي (Adobe V4) للتحويل العربي الدقيق
# ───────────────────────────────────────────────────────────────────────
def process_pdf_adobe_v4(input_path: str, output_path: str, target_format):
    client_id = os.getenv("ADOBE_CLIENT_ID")
    client_secret = os.getenv("ADOBE_CLIENT_SECRET")

    if not client_id or not client_secret:
        raise Exception("مفاتيح أدوبي غير موجودة في بيئة Railway")

    credentials = ServicePrincipalCredentials(client_id=client_id, client_secret=client_secret)
    pdf_services = PDFServices(credentials=credentials)

    with open(input_path, 'rb') as f:
        input_stream = f.read()
    input_asset = pdf_services.upload(input_stream=input_stream, mime_type=PDFServicesMediaType.PDF)

    export_pdf_params = ExportPDFParams(target_format=target_format)
    export_pdf_job = ExportPDFJob(input_asset=input_asset, export_pdf_params=export_pdf_params)

    location = pdf_services.submit(export_pdf_job)
    pdf_services_response = pdf_services.get_job_result(location, ExportPDFResult)

    result_asset = pdf_services_response.get_result().get_asset()
    stream_asset = pdf_services.get_content(result_asset)

    with open(output_path, "wb") as output_file:
        output_file.write(stream_asset.get_input_stream())

# ───────────────────────────────────────────────────────────────────────
# 2. Health + secure download (GET + HEAD)
# ───────────────────────────────────────────────────────────────────────
@app.get("/")
def health_check():
    return {"status": "PDFMasry API is running with ALL Server Tools!", "version": app.version}


@app.get("/health")
def health():
    return {"status": "ok"}


# ───────────────────────────────────────────────────────────────────────
# Admin: cache + Adobe usage stats
# ───────────────────────────────────────────────────────────────────────
# Behind a shared secret env var. No user data, no filenames — just
# counters. Fail closed: if ADMIN_STATS_TOKEN is unset, endpoint 401s
# for everyone.

@app.get("/api/admin/cache-stats")
def admin_cache_stats(request: Request):
    """Returns cache hit rates + Adobe monthly usage. Requires admin token
    passed either as ``?token=`` query param or ``X-Admin-Token`` header.

    Response shape is documented in cache_stats.snapshot(). Safe to poll
    from a dashboard; O(1) work regardless of cache size.
    """
    if not ADMIN_STATS_TOKEN:
        raise HTTPException(status_code=401, detail="Admin stats endpoint not configured")
    supplied = request.query_params.get("token") or request.headers.get("X-Admin-Token", "")
    if not hmac.compare_digest(supplied, ADMIN_STATS_TOKEN):
        raise HTTPException(status_code=401, detail="Invalid admin token")
    # cache_health removed — the file cache was retired (privacy fix).
    # Kept the endpoint shape as much as possible for backward compat.
    return {
        "cache_enabled": CACHE_ENABLED,
        "adobe_provider_version": ADOBE_PROVIDER_VERSION,
        "counters": cache_stats.snapshot(monthly_quota=ADOBE_MONTHLY_QUOTA),
        "cache_health": {"disabled": True, "entry_count": 0, "total_bytes": 0},
    }


def _build_download_headers(job_id: str, token: str) -> tuple[dict, str, str, str] | JSONResponse:
    """Shared validation + header prep for both GET and HEAD download.
    Returns (headers, media_type, download_name, output_path) on success,
    or a JSONResponse to send directly on error.
    """
    if not re.fullmatch(r"[a-f0-9]{32}", job_id):
        return JSONResponse({"detail": "Not found"}, status_code=404)
    if not _verify_token(job_id, token):
        return JSONResponse({"detail": "Invalid or expired link"}, status_code=403)

    jdir = _job_dir(job_id)
    output = os.path.join(jdir, "output.bin")
    meta_path = os.path.join(jdir, "meta.txt")
    if not os.path.isfile(output) or not os.path.isfile(meta_path):
        return JSONResponse({"detail": "File not found or already deleted"}, status_code=404)

    with open(meta_path, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()
    media_type = lines[0] if lines else "application/octet-stream"
    download_name = lines[1] if len(lines) > 1 else "download"

    headers = {
        "Content-Disposition": _content_disposition(download_name),
        "Cache-Control": "private, no-store",
        "Pragma": "no-cache",
        "X-Content-Type-Options": "nosniff",
        "Referrer-Policy": "no-referrer",
    }
    return headers, media_type, download_name, output


@app.get("/api/download/{job_id}/{token}")
@limiter.limit("60/minute")
def download_file(job_id: str, token: str, request: Request):
    result = _build_download_headers(job_id, token)
    if isinstance(result, JSONResponse):
        return result
    headers, media_type, _download_name, output = result
    response = FileResponse(output, media_type=media_type)
    for k, v in headers.items():
        response.headers[k] = v
    return response


@app.head("/api/download/{job_id}/{token}")
@limiter.limit("60/minute")
def head_download_file(job_id: str, token: str, request: Request):
    """HEAD variant: same headers as GET, no body.
    Some clients (link previewers, crawlers, download managers) probe with
    HEAD before GET. Without this handler they got 405 Method Not Allowed.
    """
    result = _build_download_headers(job_id, token)
    if isinstance(result, JSONResponse):
        return result
    headers, media_type, _download_name, output = result
    # Report actual file size so range requests and progress bars work.
    try:
        size = os.path.getsize(output)
    except OSError:
        size = 0
    headers["Content-Length"] = str(size)
    headers["Content-Type"] = media_type
    return Response(status_code=200, headers=headers)


# ───────────────────────────────────────────────────────────────────────
# 3. مسارات تحويل PDF → Office (اللغة العربية) — accept PDF only
# ───────────────────────────────────────────────────────────────────────
# Provider routing (2026-08 emergency fix):
#   Adobe PDF Services' free monthly quota (500 tx) got exhausted in
#   production, so every user request was returning 500 from the SDK.
#   We now route through LibreOffice by default — self-hosted, unlimited,
#   already on the Railway image (used for Word→PDF the other direction).
#   Adobe path is kept as an opt-in via USE_ADOBE=true env var for the
#   day the paid tier is turned on; the cache layer still fronts either
#   provider transparently.
#
# Quality note: LibreOffice's PDF→Office conversion is generally
# reasonable but weaker than Adobe on complex layouts (multi-column,
# heavy embedded fonts). For our Arabic user base, the RTL text is
# preserved correctly which is the main constraint.

USE_ADOBE = os.environ.get("USE_ADOBE", "false").lower() == "true"
LIBREOFFICE_PROVIDER_VERSION = "libreoffice-cli-2026-08"


def process_pdf_to_docx(input_path: str, output_path: str) -> None:
    """PDF → DOCX via pdf2docx (pymupdf-backed).

    Chosen over LibreOffice for two reasons:
      1. LibreOffice imports PDFs as Draw docs, so --convert-to docx just
         drops the content and produces an empty Writer file.
      2. pdf2docx preserves text runs, tables, images, and — critically for
         us — RTL text direction on Arabic content.

    Raises HTTPException 500 with a friendly Arabic message on failure.
    """
    try:
        # Lazy import so a broken install of pdf2docx doesn't crash boot.
        from pdf2docx import Converter
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة التحويل غير مثبتة على الخادم")

    try:
        cv = Converter(input_path)
        try:
            cv.convert(output_path)
        finally:
            cv.close()
    except Exception as e:
        _logger.exception("pdf→docx failed: %s", type(e).__name__)
        raise HTTPException(
            status_code=500,
            detail="تعذّر تحويل الملف إلى Word — ربما الملف تالف أو مشفَّر.",
        )

    if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
        raise HTTPException(status_code=500, detail="فشل التحويل: لم يُنشأ ملف Word صالح")


def process_pdf_to_xlsx(input_path: str, output_path: str) -> None:
    """PDF → XLSX via pdfplumber (table extraction) + openpyxl (writer).

    Each PDF page becomes its own sheet. Tables detected by pdfplumber are
    written as-is; if a page has no detectable tables, its extracted text
    is written as a single-column fallback so the user still gets content
    instead of a blank sheet.

    Preserves Arabic text (pdfplumber returns Unicode). No RTL formatting
    is applied at the cell level — Excel handles that from the character
    direction bit.
    """
    try:
        import pdfplumber
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة التحويل غير مثبتة على الخادم")

    try:
        wb = Workbook()
        # Remove default sheet — we'll add one per page
        default_sheet = wb.active
        wb.remove(default_sheet)

        with pdfplumber.open(input_path) as pdf:
            if not pdf.pages:
                raise HTTPException(status_code=500, detail="الملف لا يحتوي على صفحات")

            for page_num, page in enumerate(pdf.pages, start=1):
                sheet_name = f"Page {page_num}"[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=sheet_name)
                tables = page.extract_tables() or []
                if tables:
                    for table in tables:
                        for row in table:
                            # openpyxl expects a list; None cells become empty
                            ws.append([cell if cell is not None else "" for cell in row])
                        ws.append([])  # blank row between tables
                else:
                    # No tables detected — fall back to line-by-line text so
                    # the user isn't left with an empty sheet.
                    text = page.extract_text() or ""
                    for line in text.splitlines():
                        if line.strip():
                            ws.append([line])

        # Guard against a completely empty workbook (all pages truly blank)
        if not wb.sheetnames:
            wb.create_sheet(title="Empty")

        wb.save(output_path)
    except HTTPException:
        raise
    except Exception as e:
        _logger.exception("pdf→xlsx failed: %s", type(e).__name__)
        raise HTTPException(
            status_code=500,
            detail="تعذّر تحويل الملف إلى Excel.",
        )

    if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
        raise HTTPException(status_code=500, detail="فشل التحويل: لم يُنشأ ملف Excel صالح")


async def _convert_pdf_to_office(
    endpoint: str,
    input_path: str,
    output_path: str,
    target_ext: str,
    adobe_format,
) -> None:
    """Run PDF→Office conversion with provider auto-selection. No caching.

    Args:
        endpoint: tool name for stats attribution ("pdf-to-word" / "pdf-to-excel")
        input_path: user-uploaded PDF
        output_path: destination for the DOCX/XLSX result
        target_ext: "docx" or "xlsx"
        adobe_format: Adobe ExportPDFTargetFormat enum (only used when
            USE_ADOBE is enabled)

    Provider precedence:
        1. USE_ADOBE=true → Adobe SDK
        2. otherwise → pdf2docx / pdfplumber (self-hosted)

    File cache was removed to match the site's published 1-hour deletion
    promise. Every request pays the full conversion cost; that's fine —
    pdf2docx typically runs in 0.5-2s per document.
    """
    if USE_ADOBE:
        await asyncio.to_thread(process_pdf_adobe_v4, input_path, output_path, adobe_format)
        cache_stats.record_adobe_transaction(endpoint)
    elif target_ext == "docx":
        # to_thread keeps FastAPI's event loop responsive while pdf2docx
        # crunches PyMuPDF-heavy work in the background thread pool.
        await asyncio.to_thread(process_pdf_to_docx, input_path, output_path)
    elif target_ext == "xlsx":
        await asyncio.to_thread(process_pdf_to_xlsx, input_path, output_path)
    else:
        raise HTTPException(status_code=500, detail=f"صيغة إخراج غير مدعومة: {target_ext}")


@app.post("/api/pdf-to-word")
@limiter.limit("20/minute")
async def convert_pdf_to_word(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    job_id, jdir, orig = await _save_upload_to_job(file, {"pdf"})
    input_path = os.path.join(jdir, "input.bin")
    output_path = os.path.join(jdir, "converted.docx")
    try:
        await _convert_pdf_to_office("pdf-to-word", input_path, output_path, "docx", ExportPDFTargetFormat.DOCX)
        base = os.path.splitext(orig)[0] or "document"
        response_data = _make_response(
            job_id,
            output_path,
            f"{base}.docx",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except HTTPException:
        # Preserve the tailored status + Arabic detail already raised by the provider.
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise
    except Exception as e:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        _logger.exception("pdf-to-word endpoint: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="تعذّر تحويل الملف. يرجى المحاولة مرة أخرى.")


@app.post("/api/pdf-to-excel")
@limiter.limit("20/minute")
async def convert_pdf_to_excel(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    job_id, jdir, orig = await _save_upload_to_job(file, {"pdf"})
    input_path = os.path.join(jdir, "input.bin")
    output_path = os.path.join(jdir, "converted.xlsx")
    try:
        await _convert_pdf_to_office("pdf-to-excel", input_path, output_path, "xlsx", ExportPDFTargetFormat.XLSX)
        base = os.path.splitext(orig)[0] or "document"
        response_data = _make_response(
            job_id,
            output_path,
            f"{base}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except HTTPException:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise
    except Exception as e:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        _logger.exception("pdf-to-excel endpoint: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="تعذّر تحويل الملف. يرجى المحاولة مرة أخرى.")

# ───────────────────────────────────────────────────────────────────────
# 4. الأدوات المجانية (Ghostscript, qpdf, LibreOffice, Poppler) — PDF
# ───────────────────────────────────────────────────────────────────────
@app.post("/api/compress")
@limiter.limit("20/minute")
async def compress_pdf(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    job_id, jdir, orig = await _save_upload_to_job(file, {"pdf"})
    input_path = os.path.join(jdir, "input.bin")
    output_path = os.path.join(jdir, "compressed.pdf")
    try:
        cmd = ["gs", "-sDEVICE=pdfwrite", "-dCompatibilityLevel=1.4",
               "-dPDFSETTINGS=/screen", "-dNOPAUSE", "-dQUIET", "-dBATCH",
               f"-sOutputFile={output_path}", input_path]
        await _run_subprocess(cmd, timeout=60)
        response_data = _make_response(job_id, output_path, f"compressed_{orig}", "application/pdf")
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except Exception:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise HTTPException(status_code=500, detail="فشل ضغط الملف")


@app.post("/api/protect")
@limiter.limit("15/minute")
async def protect_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    password: str = Form("pdfmasry"),
):
    job_id, jdir, orig = await _save_upload_to_job(file, {"pdf"})
    input_path = os.path.join(jdir, "input.bin")
    output_path = os.path.join(jdir, "protected.pdf")
    try:
        cmd = ["qpdf", "--encrypt", password, password, "256", "--", input_path, output_path]
        await _run_subprocess(cmd, timeout=45)
        response_data = _make_response(job_id, output_path, f"protected_{orig}", "application/pdf")
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except Exception:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise HTTPException(status_code=500, detail="فشلت حماية الملف")


@app.post("/api/unlock")
@limiter.limit("15/minute")
async def unlock_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    password: str = Form("pdfmasry"),
):
    job_id, jdir, orig = await _save_upload_to_job(file, {"pdf"})
    input_path = os.path.join(jdir, "input.bin")
    output_path = os.path.join(jdir, "unlocked.pdf")
    try:
        cmd = ["qpdf", f"--password={password}", "--decrypt", input_path, output_path]
        await _run_subprocess(cmd, timeout=45)
        response_data = _make_response(job_id, output_path, f"unlocked_{orig}", "application/pdf")
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except Exception:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise HTTPException(status_code=500, detail="فشل فك الحماية، قد تكون كلمة المرور خاطئة")


@app.post("/api/pdf-to-image")
@limiter.limit("20/minute")
async def pdf_to_image(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    job_id, jdir, orig = await _save_upload_to_job(file, {"pdf"})
    input_path = os.path.join(jdir, "input.bin")
    pages_dir = os.path.join(jdir, "pages")
    os.makedirs(pages_dir, exist_ok=True)
    try:
        cmd = ["pdftoppm", "-jpeg", "-r", "150", input_path, os.path.join(pages_dir, "page")]
        await _run_subprocess(cmd, timeout=90)

        zip_base = os.path.join(jdir, "pages_archive")
        shutil.make_archive(zip_base, "zip", pages_dir)
        zip_path = f"{zip_base}.zip"

        base = os.path.splitext(orig)[0] or "document"
        response_data = _make_response(job_id, zip_path, f"{base}_images.zip", "application/zip")
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except Exception:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise HTTPException(status_code=500, detail="فشل تحويل الملف إلى صور")


@app.post("/api/word-to-pdf")
@app.post("/api/excel-to-pdf")
@limiter.limit("20/minute")
async def office_to_pdf(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    # Accept both Office 2007+ (zip) and legacy 97-2003 (ole).
    job_id, jdir, orig = await _save_upload_to_job(file, {"office_zip", "office_ole"})
    # LibreOffice needs the input to have a proper extension
    input_ext = os.path.splitext(orig)[1] or ".docx"
    input_path = os.path.join(jdir, f"input{input_ext}")
    canonical_input = os.path.join(jdir, "input.bin")
    shutil.move(canonical_input, input_path)

    try:
        cmd = ["libreoffice", "--headless", "--convert-to", "pdf", input_path, "--outdir", jdir]
        await _run_subprocess(cmd, timeout=120)
        # LibreOffice writes <basename>.pdf where basename = os.path.splitext(input_filename)[0]
        libre_output = os.path.join(jdir, f"input.pdf")
        if not os.path.exists(libre_output):
            # Some versions name it after original — glob for any .pdf just in case
            candidates = [f for f in os.listdir(jdir) if f.lower().endswith(".pdf")]
            if candidates:
                libre_output = os.path.join(jdir, candidates[0])
            else:
                raise RuntimeError("LibreOffice produced no PDF")

        base = os.path.splitext(orig)[0] or "document"
        response_data = _make_response(job_id, libre_output, f"{base}.pdf", "application/pdf")
        background_tasks.add_task(_delete_job_after_delay, job_id)
        return response_data
    except Exception:
        background_tasks.add_task(_delete_job_after_delay, job_id, delay_seconds=5)
        raise HTTPException(status_code=500, detail="فشل تحويل المستند إلى PDF")
